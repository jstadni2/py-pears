import os
import pandas as pd
from faker import Faker
from faker_education import SchoolProvider
import random
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side
import shutil

import py_pears.utils as utils


# Calculate the path to the root directory of this package
ROOT_DIR = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))

PY_PEARS_DIR = ROOT_DIR + '/py_pears'  # not necessary?
EXPORT_DIR = PY_PEARS_DIR + '/pears_exports/'

TEST_INPUTS_DIR = ROOT_DIR + '/tests/test_inputs/'
TEST_INPUTS_PEARS_DIR = TEST_INPUTS_DIR + 'pears/'
TEST_INPUTS_PEARS_PREV_YEAR_DIR = TEST_INPUTS_PEARS_DIR + 'prev_year/'
TEST_COALITION_SURVEY_EXPORTS_DIR = TEST_INPUTS_PEARS_DIR + 'coalition_survey_exports/'

# Faker settings
fake = Faker()
fake.add_provider(SchoolProvider)
# set seed

# Functions for generating fake data


def create_user(num=1):
    user = [{"full_name": fake.name()} for x in range(num)]
    return user


# Use interface for duplicated code across similar functions?


# generate random latitude between 37.720129 through 41.92947
def random_lat(start=37.720129, stop=41.92947):
    return random.uniform(start, stop)


# generate random longitude -91.407676 through -87.591379
def random_long(start=-91.407676, stop=-87.591379):
    return random.uniform(start, stop)


def create_site(num=1):
    user = [{'site_name': fake.company(),
             'address': fake.street_address(),
             'latitude': random_lat(),
             'longitude': random_long()} for x in range(num)]
    return user


def create_school(num=1):
    school = [{'site_name': fake.school_name(),
               'address': fake.street_address(),
               'latitude': random_lat(),
               'longitude': random_long()} for x in range(num)]
    return school


def create_district(num=1):
    district = [{'site_name': fake.school_district(),
                 'address': fake.street_address(),
                 'latitude': random_lat(),
                 'longitude': random_long()} for x in range(num)]
    return district


def join_fake_sites(sites, site_fields, site_type):
    fake_sites = sites.copy().drop(site_fields, axis=1).reset_index()
    num_sites = len(fake_sites)
    fake_dict = dict()

    if site_type == 'site':
        fake_dict = create_site(num_sites)
    elif site_type == 'school':
        fake_dict = create_school(num_sites)
    elif site_type == 'district':
        fake_dict = create_district(num_sites)

    fake_sites = fake_sites.join(pd.DataFrame(fake_dict)).drop('index', axis=1)

    for field in site_fields:
        fake_sites.insert(sites.columns.get_loc(field), field, fake_sites.pop(field))

    return fake_sites


# Functions for Excel Workbook manipulation
# Move to utils.py?


# Helper function to replace Excel sheet with dataframe
def overwrite_sheet(file_name, book, sheet, df):
    if sheet not in book.sheetnames:
        raise Exception('sheet: ' + sheet + ' not found in Excel Workbook: ' + file_name)

    ws = book[sheet]
    ws.delete_cols(1, ws.max_column)
    ws.delete_rows(1, ws.max_row)

    rows = dataframe_to_rows(df, index=False)
    for row in rows:
        ws.append(row)

    # Use original sheet's style formatting
    font = Font(bold=False)
    fill = PatternFill(fill_type='solid', fgColor='00C0C0C0')
    border = Border(bottom=Side(border_style=None, color='FF000000'))
    for c in ws["1:1"]:
        c.font = font
        c.fill = fill
        c.border = border

    book.save(file_name)


# Function to update workbook with reformatted data
# file_name: Module export Excel file
# sheets: list of Excel sheets
# data: list of dataframes to overwrite sheet data
# sheets_dict
def overwrite_excel(file_name, sheets, data):
    sheets_dict = dict(zip(sheets, data))
    book = openpyxl.load_workbook(file_name)
    for sheet, df in sheets_dict.items():
        overwrite_sheet(file_name, book, sheet, df)


def delete_sheets(file_name, sheets):
    wb = openpyxl.load_workbook(file_name)
    for sheet in sheets:
        if sheet in wb.sheetnames:
            wb.remove(wb[sheet])
    wb.save(file_name)


# Utilities for replacing data

def replace_dict(old_values, new_values):
    return dict(zip(old_values, new_values))


# substitutes dataframe column by merging in fake data
def sub_fake_data(df, col, fake_data, merge_on):
    df_out = df.copy()
    idx = df_out.columns.get_loc(col)
    df_out = pd.merge(df_out.drop(col, axis=1), fake_data, how='left', on=merge_on)
    df_out.insert(idx, col, df_out.pop(col))
    return df_out


# Randomize specific module metrics


# function for making dummy text values for column
def randomize_text(module, df, module_id, col):
    out_df = df.copy()
    out_df[col] = module + ' ' + out_df[module_id].map(str) + ' ' + col.replace('_', ' ').title()
    return out_df


# function for randomizing numeric column values
def randomize_metric(df, metric):
    out_df = df.copy()
    fake_dict = [{metric: random.randint(out_df[metric].min(),
                                         out_df[metric].max())} for x in range(len(out_df))]
    out_df[metric] = pd.DataFrame(fake_dict)
    return out_df


# Class for navigating PEARS modules
class Module:
    def __init__(self, name, submodules):
        self.name = name
        self.submodules = submodules


# Class for navigating PEARS submodules
class Submodule:
    def __init__(self, name,
                 user_fields=[],
                 site_fields=[],
                 text_fields=[],
                 numeric_fields=[]):
        self.name = name
        self.user_fields = user_fields
        self.site_fields = site_fields
        self.text_fields = text_fields
        self.numeric_fields = numeric_fields


# Replace data in PEARS Module workbooks with fake data
# in_path
# out_path
# import_modules: list of Module objects
# emails_dict
# names_dict
# fake_sites
def clean_module_exports(in_path, out_path, import_modules, emails_dict, names_dict, fake_sites):
    for module in import_modules:
        src = in_path + module.name + "_Export.xlsx"
        dst = out_path + module.name + "_Export.xlsx"
        shutil.copyfile(src, dst)
        # ADD open book, instantiate writer
        book = openpyxl.load_workbook(dst)
        # Read module's sheet
        for submod in module.submodules:
            data = pd.read_excel(dst, sheet_name=submod.name)

            for field in submod.user_fields:
                if field == 'reported_by_email':
                    data[field] = data[field].replace(emails_dict, regex=True)
                else:
                    data[field] = data[field].replace({'\(': '', '\)': ''}, regex=True).replace(names_dict, regex=True)

            for field in submod.site_fields:
                data = sub_fake_data(data, field, fake_sites[['site_id', field]], 'site_id')

            for field in submod.text_fields:
                data = randomize_text(submod.name, data, data.columns[0], field)

            for field in submod.numeric_fields:
                data = randomize_metric(data, field)

            # Overwrite each workbook sheet with replaced data
            # use overwrite_sheet instead
            overwrite_sheet(dst, book, submod.name, data)


# Function to remove worksheet format
# ws: worksheet object
def remove_format(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.style = 'Normal'
            # cell.fill = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
            cell.font = Font(bold=False)
            cell.border = Border(bottom=Side(border_style=None, color='FF000000'))


def clean_staff_list(test_staff_list, emails_dict, netids_dict, users,
                     last_names_dict, first_names_dict, last_first_dict):

    staff_wb = openpyxl.load_workbook(test_staff_list)

    # Iterate through all sheets
    for sheet in staff_wb.sheetnames:
        ws = staff_wb[sheet]
        # Unlock sheet
        ws.protection.disable()

        data = ws.values

        # Set column labels based on sheet name
        columns = next(data)[0:]
        if sheet in ['SNAP-Ed Staff List', 'HEAT Project Staff', 'FCS State Office', 'ISBE Staff List',
                     'EFNEP Staff List']:
            columns = next(data)[0:]
        elif sheet == 'CPHP Staff List':
            while columns[0] != 'Last Name':
                columns = next(data)[0:]

        df = pd.DataFrame(data, columns=columns)
        df = df[df.columns.drop(list(df.filter(regex='Column')))]

        # Rename columns for specific sheets
        if sheet == "RE's and CD's":
            df = df.rename(columns={'NETID/E-MAIL': 'RE E-MAIL',
                                    'E-MAIL': 'CD E-MAIL'})
        if sheet == 'Former Staff':
            df = df.rename(columns={'E-MAIL/NETID': 'NETID'})

        # Iterate through all fields
        for col in df.columns:
            if col is None:
                continue
            elif col == 'MISC/NOTES' or 'Phone' in col:
                df = df.drop(columns=col)
            elif 'E-MAIL' in col or col == 'Email Address':
                df[col] = df[col].replace(emails_dict, regex=True)
                df = df.loc[df[col].str.contains('@fake_domain.com', na=False)]
            elif col == 'NETID':
                df[col] = df[col].replace(netids_dict, regex=True)
                df = df.loc[df[col].isin(users['new_last_name'])]
            elif col == 'Last Name':
                df[col] = df[col].replace(last_names_dict, regex=True)
                df = df.loc[df[col].isin(users['new_last_name'])]
            elif col == 'First Name':
                df[col] = df[col].replace(first_names_dict, regex=True)
                # df = df.loc[df[col] == 'User']
            elif col in ['NAME', 'REGIONAL EDUCATOR', 'COUNTY DIRECTOR']:
                df[col] = df[col].replace(last_first_dict, regex=True)
                # if col == 'NAME':
                df = df.loc[df[col].isin(users['new_last_first'])]

        # Remove old sheet, create new one with updated data
        staff_wb.remove(ws)
        staff_wb.create_sheet(sheet)
        ws = staff_wb[sheet]

        rows = dataframe_to_rows(df, index=False)
        for row in rows:
            ws.append(row)

    staff_wb.save(test_staff_list)


# test_pears_dir: the target directory for PEARS imports
# test_inputs_dir=TEST_INPUTS_DIR
# staff_src
# pears_prev_year_dir
# test_inputs_pears_prev_year_dir
# test_coalition_surveys_dir
def main(export_dir=EXPORT_DIR, test_inputs_dir=TEST_INPUTS_DIR, test_pears_dir=TEST_INPUTS_PEARS_DIR,
         test_coalition_surveys_dir=TEST_COALITION_SURVEY_EXPORTS_DIR):

    creds = utils.load_credentials()

    # Download all PEARS S3 objects for today
    utils.download_s3_exports(profile=creds['aws_profile'],
                              org=creds['s3_organization'],
                              dst=export_dir)

    # Import Users export
    users = pd.read_excel(export_dir + 'User_Export.xlsx',
                          sheet_name='User Data')[
        ['user_id', 'username', 'email', 'full_name', 'unit', 'program_area', 'viewable_units', 'is_active']]

    # Set replacement values for users

    users['new_full_name'] = pd.DataFrame(create_user(len(users)))
    users['new_last_name'] = users['new_full_name'].str.split(pat=' ', n=1).str[1]
    users['new_first_name'] = users['new_full_name'].str.split(pat=' ', n=1).str[0]
    users['new_last_first'] = users['new_last_name'] + ',' + ' ' + users['new_first_name']
    users['new_username'] = users['new_first_name'].str.replace('.', '', regex=True) + '.' + users[
        'new_last_name'].str.replace('.', '', regex=True) + '@fake_domain.com'
    users['new_username'] = users['new_username'].replace({' ': '.'}, regex=True)
    users['new_email'] = users['new_username']

    # Create replace dicts for site_user_fields

    emails_dict = replace_dict(users['email'], users['new_email'])
    names_dict = replace_dict(users['full_name'].replace({'\(': '', '\)': ''}, regex=True),
                              users['new_full_name'])

    # Clean PEARS users export

    cleaned_users = users[
        ['user_id', 'new_username', 'new_email', 'new_full_name', 'unit', 'program_area', 'viewable_units',
         'is_active']]
    cleaned_users.columns = cleaned_users.columns.str.replace("new_", "")

    cleaned_users_filename = test_pears_dir + 'User_Export.xlsx'

    shutil.copyfile(export_dir + 'User_Export.xlsx', cleaned_users_filename)
    overwrite_excel(cleaned_users_filename, ['User Data'], [cleaned_users])

    # Delete other tabs
    delete_sheets(cleaned_users_filename, ['Program Area Team Members', 'Quarterly Effort Report Checkup'])

    # Clean sites export

    sites_src = export_dir + 'Site_Export.xlsx'
    sites_dst = test_pears_dir + 'Site_Export.xlsx'
    shutil.copyfile(sites_src, sites_dst)

    cleaned_sites = pd.read_excel(sites_src, sheet_name='Site Data')

    site_user_fields = ['created_by', 'created_by_email', 'modified_by']

    for field in site_user_fields:
        if field == 'created_by_email':
            cleaned_sites[field] = cleaned_sites[field].replace(emails_dict, regex=True)
        else:
            cleaned_sites[field] = cleaned_sites[field].replace({'\(': '', '\)': ''}, regex=True).replace(names_dict,
                                                                                                          regex=True)

    cleaned_sites = cleaned_sites.loc[~cleaned_sites['created_by_email'].str.contains('canopyteam', na=False)]

    cleaned_sites.loc[~cleaned_sites['modified_by'].isin(users['new_full_name']) &
                      ((cleaned_sites['modified_by'] != '') | cleaned_sites[
                          'modified_by'].notnull()), 'modified_by'] = ''

    contact_fields = ['contact_name', 'contact_email', 'contact_phone']

    for field in contact_fields:
        cleaned_sites[field] = ''

    site_fields = ['site_name', 'address', 'latitude', 'longitude']

    # Create fake data for site_fields

    # Subset schools
    cleaned_schools = cleaned_sites.loc[~(cleaned_sites['site_name'].str.contains('District', na=False)) &
                                        (cleaned_sites['setting'] == 'Schools (K-12, elementary, middle, and high)')]

    cleaned_schools = join_fake_sites(cleaned_schools, site_fields, 'school')

    # Subset districts
    cleaned_districts = cleaned_sites.loc[(cleaned_sites['site_name'].str.contains('District', na=False)) &
                                          (cleaned_sites['setting'] == 'Schools (K-12, elementary, middle, and high)')]

    cleaned_districts = join_fake_sites(cleaned_districts, site_fields, 'district')

    # Subset all other sites
    cleaned_sites = cleaned_sites.loc[~cleaned_sites['site_id'].isin(cleaned_schools['site_id']) &
                                      ~cleaned_sites['site_id'].isin(cleaned_districts['site_id'])]

    cleaned_sites = join_fake_sites(cleaned_sites, site_fields, 'site')

    cleaned_sites = pd.concat([cleaned_sites, cleaned_schools, cleaned_districts])

    # merge parent_site_name using parent_site_id
    cleaned_sites = sub_fake_data(cleaned_sites,
                                  'parent_site_name',
                                  cleaned_sites[['site_id',
                                                 'site_name']].rename(columns={'site_id': 'parent_site_id',
                                                                               'site_name': 'parent_site_name'}),
                                  'parent_site_id')

    cleaned_sites['comment'] = ''

    # Demographics tab, 'site_name'

    cleaned_demo = pd.read_excel(sites_src, sheet_name='Demographics')

    cleaned_demo = sub_fake_data(cleaned_demo, 'site_name', cleaned_sites[['site_id', 'site_name']], 'site_id')

    overwrite_excel(sites_dst, ['Site Data', 'Demographics'], [cleaned_sites, cleaned_demo])

    # Clean import modules

    collaborators = Submodule('Collaborators', user_fields=['user'], site_fields=[])

    # Default field lists
    user_fields = ['reported_by', 'reported_by_email', 'updated_by', 'collaborators']
    site_fields = ['site_name', 'site_address', 'site_latitude', 'site_longitude']

    program_activities = Module('Program_Activities',
                                [
                                    Submodule('Program Activity Data',
                                              user_fields=['reported_by', 'reported_by_email', 'collaborators',
                                                           'contributors'],
                                              site_fields=site_fields,
                                              text_fields=['name', 'copied_from', 'comments'],
                                              numeric_fields=['participants_total', 'number_sessions']),
                                    Submodule('Sessions', numeric_fields=['num_participants']),
                                    collaborators
                                ])

    indirect_activities = Module('Indirect_Activity',
                                 [
                                     Submodule('Indirect Activity Data',
                                               user_fields=user_fields,
                                               text_fields=['comments', 'title']),
                                     Submodule('Intervention Channels',
                                               site_fields=site_fields,
                                               numeric_fields=['reach']),
                                     collaborators
                                 ])

    partnerships = Module('Partnership',
                          [
                              Submodule('Partnership Data',
                                        user_fields=user_fields,
                                        site_fields=site_fields,
                                        text_fields=['partnership_name', 'copied_from', 'comment', 'accomplishments',
                                                     'lessons_learned']),
                              Submodule('Meetings',
                                        text_fields=['event_title', 'notes'],
                                        numeric_fields=['attendance']),
                              collaborators
                          ])

    coalitions = Module('Coalition',
                        [
                            Submodule('Coalition Data',
                                      user_fields=user_fields,
                                      text_fields=['coalition_name', 'event_title', 'comment', 'accomplishments',
                                                   'copied_from'],
                                      numeric_fields=['number_of_members']),
                            Submodule('Members', site_fields=site_fields, text_fields=['name', 'role_and_resources']),
                            Submodule('Meetings',
                                      text_fields=['event_title', 'notes'],
                                      numeric_fields=['attendance']),
                            collaborators
                        ])

    pse_site_activities = Module('PSE_Site_Activity',
                                 [
                                     Submodule('PSE Data',
                                               user_fields=['reported_by', 'reported_by_email', 'collaborators',
                                                            'contributors'],
                                               site_fields=['site_name', 'site_address'],
                                               text_fields=['name', 'copied_from', 'comments', 'contribution_notes',
                                                            'other_changes_made',
                                                            'additional_barriers_or_facilitators',
                                                            'influence_on_future_pse_work',
                                                            'improvements_or_changes'],
                                               numeric_fields=['total_reach']),
                                     Submodule('Needs, Readiness, Effectiveness',
                                               text_fields=['baseline_results', 'follow_up_results']),
                                     Submodule('Recognition and Media Coverage',
                                               text_fields=['name', 'recognizing_body', 'link_or_reference']),
                                     collaborators
                                 ])

    success_stories = Module('Success_Story',
                             [
                                 Submodule('Success Story Data',
                                           user_fields=user_fields,
                                           site_fields=['site_name', 'site_address'],
                                           text_fields=['title', 'comments', 'background', 'story_narrative',
                                                        'favorite_quote', 'program_activity']),
                                 collaborators
                             ])

    import_modules = [program_activities, indirect_activities, partnerships, coalitions, pse_site_activities,
                      success_stories]

    clean_module_exports(in_path=export_dir,
                         out_path=test_pears_dir,
                         import_modules=import_modules,
                         emails_dict=emails_dict,
                         names_dict=names_dict,
                         fake_sites=cleaned_sites.rename(columns={'address': 'site_address',
                                                                  'latitude': 'site_latitude',
                                                                  'longitude': 'site_longitude'})[
                             ['site_id', 'site_name', 'site_address', 'site_latitude', 'site_longitude']])

    # Clean staff list

    # Create fields for last name, first name, netid

    users['last_name'] = users['full_name'].str.split(pat=' ', n=1).str[1]
    users['first_name'] = users['full_name'].str.split(pat=' ', n=1).str[0]
    users['last_first'] = users['last_name'] + ', ' + users['first_name']
    users['netid'] = users['email'].replace({'@illinois.edu': '', '@uic.edu': ''}, regex=True)

    # Create additional dicts

    first_names_dict = replace_dict(users['first_name'], users['new_first_name'])
    last_names_dict = replace_dict(users['last_name'], users['new_last_name'])
    last_first_dict = replace_dict(users['last_first'], users['new_last_first'])
    netids_dict = replace_dict(users['netid'], users['new_last_name'])

    # Copy staff list

    staff_fp = test_inputs_dir + 'FY23_INEP_Staff_List.xlsx'
    shutil.copyfile(creds['staff_list'], staff_fp)

    clean_staff_list(staff_fp, emails_dict, netids_dict, users,
                     last_names_dict, first_names_dict, last_first_dict)

    # Clean FY 2021 PEARS Exports

    clean_module_exports(creds['pears_prev_year'],
                         test_pears_dir + 'prev_year/',
                         import_modules,
                         emails_dict,
                         names_dict,
                         cleaned_sites.rename(columns={'address': 'site_address',
                                                       'latitude': 'site_latitude',
                                                       'longitude': 'site_longitude'})[
                             ['site_id', 'site_name', 'site_address', 'site_latitude', 'site_longitude']])

    # Clean PEARS Coalition Survey Exports

    coalition_surveys = [Module('Coalition_Survey_' + q,
                                [
                                    Submodule('Response Data',
                                              user_fields=['Program Entered By'],
                                              text_fields=['Program Name', 'coalition_name',
                                                           'Please list any other goals your coalition is working '
                                                           'towards which do not fit under any of the goals listed '
                                                           'above.',
                                                           # Since the following fields are manually entered, some
                                                           # values remain when input as user_fields
                                                           'staff_name',
                                                           'staff_email'])
                                    # missed my user_fields cleaning because label isn't 'reported_by_email'
                                ]) for q in ['Q1', 'Q2', 'Q3', 'Q4']]

    # Create a separate function for cleaning/generating survey response data?
    # make optional arg
    clean_module_exports(creds['coalition_survey_exports'],
                         test_coalition_surveys_dir,
                         coalition_surveys, emails_dict,
                         names_dict,
                         cleaned_sites.rename(columns={'address': 'site_address',
                                                       'latitude': 'site_latitude',
                                                       'longitude': 'site_longitude'})[
                             ['site_id', 'site_name', 'site_address', 'site_latitude', 'site_longitude']]
                         )


if __name__ == '__main__':
    main()
