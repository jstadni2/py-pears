import pytest

import pandas as pd
import boto3
from botocore.exceptions import ClientError
import os
import openpyxl

import py_pears.utils as utils


# Test PEARS AWS S3

creds = utils.load_org_settings()
profile = creds['aws_profile']
org = creds['s3_organization']
my_bucket = 'exports.pears.oeie.org'
date = pd.to_datetime("today").strftime("%Y/%m/%d")


def test_aws_profile():
    try:
        boto3.Session(profile_name=profile)
    except Exception as e:
        assert False, f"'aws_profile' in org_settings.json raised an exception: \n{e}"


def test_pears_bucket():
    try:
        session = boto3.Session(profile_name=profile)
        conn = session.client('s3')

        conn.list_objects_v2(
            Bucket=my_bucket,
            Prefix=org + '/' + date + '/',
            MaxKeys=5)
    except conn.exceptions.NoSuchBucket as e:
        assert False, f"Bucket: {my_bucket} does not exist. Contact PEARS support via support@pears.io. \n{e}"
    except ClientError as e:
        assert False, f"Verify 's3_organization' in org_settings.json is set correctly. \n" \
                      f"Attempt to connect to the PEARS AWS S3 raised an exception: \n{e}"
    except Exception as e:
        assert False, f"Attempt to connect to the PEARS AWS S3 raised an exception: \n{e}"


def test_pears_s3_objects():
    session = boto3.Session(profile_name=profile)
    conn = session.client('s3')

    response = conn.list_objects_v2(
        Bucket=my_bucket,
        Prefix=org + '/' + date + '/',
        MaxKeys=20)

    assert 'Contents' in response, f"No directory or contents for {date}"

    object_filenames = []
    for f in response['Contents']:
        file = f['Key']
        object_filenames.append(file[file.rfind('/') + 1:])

    expected_exports = ['Action_Plan_Outcomes_Export.xlsx',
                        'Action_Plans_Export.xlsx',
                        'Coalition_Export.xlsx',
                        'Indirect_Activity_Export.xlsx',
                        'PSE_Site_Activity_Export.xlsx',
                        'Partnership_Export.xlsx',
                        'Program_Activities_Export.xlsx',
                        'Quarterly_Effort_Export.xlsx',
                        'Site_Export.xlsx',
                        'Social_Marketing_Campaigns_Export.xlsx',
                        'Success_Story_Export.xlsx',
                        'User_Export.xlsx']

    assert object_filenames == expected_exports


# Test PEARS Export workbook schema

# Calculate the path to the root directory of this package
ROOT_DIR = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))

ACTUAL_EXPORTS_DIR = ROOT_DIR + '/py_pears/pears_exports/'
EXPECTED_EXPORTS_DIR = ROOT_DIR + '/tests/test_inputs/pears/'

utils.download_s3_exports(profile=profile,
                          org=org,
                          dst=ACTUAL_EXPORTS_DIR)

exports = os.listdir(ACTUAL_EXPORTS_DIR)


def compare_sheets(xlsx1, xlsx2):
    sheet_names1 = openpyxl.load_workbook(xlsx1).sheetnames
    sheet_names2 = openpyxl.load_workbook(xlsx2).sheetnames
    return sheet_names1 == sheet_names2


def test_export_sheets():
    for export in exports:
        if export in ['.gitignore', 'User_Export.xlsx']:  # Remove User Export once all test input sheets are cleaned
            continue
        assert compare_sheets(ACTUAL_EXPORTS_DIR + export, EXPECTED_EXPORTS_DIR + export)


def compare_code_books(xlsx1, xlsx2):
    wb1 = openpyxl.load_workbook(xlsx1)
    wb2 = openpyxl.load_workbook(xlsx2)

    df1 = pd.DataFrame(wb1['Codebook'].values)
    df2 = pd.DataFrame(wb2['Codebook'].values)

    # Remove Exported by timestamp
    df1.iloc[4, 0] = None
    df2.iloc[4, 0] = None

    return df1.equals(df2)


def test_code_books():
    for export in exports:
        if export in ['.gitignore']:
            continue
        assert compare_code_books(ACTUAL_EXPORTS_DIR + export, EXPECTED_EXPORTS_DIR + export)


def compare_sheet_fields(xlsx1, xlsx2):
    wb1 = openpyxl.load_workbook(xlsx1)
    wb2 = openpyxl.load_workbook(xlsx2)

    for sheet in wb1.sheetnames:
        if sheet == 'Codebook':
            continue

        df1 = utils.first_row_to_cols(pd.DataFrame(wb1[sheet].values))
        df2 = utils.first_row_to_cols(pd.DataFrame(wb2[sheet].values))
        fields1 = df1.columns.tolist()
        fields2 = df2.columns.tolist()

        if fields1 != fields2:
            return False

    return True


def test_sheet_fields():
    for export in exports:
        if export in ['.gitignore', 'User_Export.xlsx']:  # Remove User Export once all test input sheets are cleaned
            continue
        assert compare_sheet_fields(ACTUAL_EXPORTS_DIR + export, EXPECTED_EXPORTS_DIR + export)
