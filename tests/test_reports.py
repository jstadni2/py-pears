import pytest

import os
import openpyxl
import pandas as pd
import numpy as np

import py_pears.utils as utils
import py_pears.reports.sites_report as sites_report
import py_pears.reports.staff_report as staff_report
import py_pears.reports.monthly_data_cleaning as monthly_data_cleaning
import py_pears.reports.quarterly_program_evaluation as quarterly_program_evaluation
import py_pears.reports.partnerships_entry as partnerships_entry
import py_pears.reports.coalition_survey_cleaning as coalition_survey_cleaning


# Calculate the path to the root directory of this package
ROOT_DIR = os.path.realpath(os.path.join(os.path.dirname(__file__), '.'))

# Set directories using package defaults
TEST_INPUTS_DIR = ROOT_DIR + '/test_inputs/'
TEST_INPUTS_PEARS_DIR = TEST_INPUTS_DIR + '/pears/'
TEST_INPUTS_PEARS_PREV_YEAR_DIR = TEST_INPUTS_PEARS_DIR + '/prev_year/'
TEST_INPUTS_PEARS_COA_SURVEYS_DIR = TEST_INPUTS_PEARS_DIR + '/coalition_survey_exports/'

# Set paths to external data inputs
staff_list = TEST_INPUTS_DIR + 'FY23_INEP_Staff_List.xlsx'
names_list = TEST_INPUTS_DIR + 'BABY_NAMES_IL.TXT'
unit_counties = TEST_INPUTS_DIR + 'Illinois Extension Unit Counties.xlsx'
update_notifications = TEST_INPUTS_DIR + 'Update Notifications.xlsx'

EXPECTED_OUTPUTS_DIR = ROOT_DIR + '/expected_outputs/'
ACTUAL_OUTPUTS_DIR = ROOT_DIR + '/actual_outputs/'

creds = utils.load_credentials()

staff_list = TEST_INPUTS_DIR + 'FY23_INEP_Staff_List.xlsx'
# Set following paths to external data inputs instead of test inputs
names_list = TEST_INPUTS_DIR + 'BABY_NAMES_IL.TXT'
unit_counties = TEST_INPUTS_DIR + 'Illinois Extension Unit Counties.xlsx'
update_notifications = TEST_INPUTS_DIR + 'Update Notifications.xlsx'

prev_month = (pd.to_datetime("today") - pd.DateOffset(months=1))


# Compare Excel Workbook objects
def compare_workbooks(xlsx1, xlsx2, diff_filename):
    wb1 = openpyxl.load_workbook(xlsx1)  # use openpyxl?
    wb2 = openpyxl.load_workbook(xlsx2)
    # Return False if sheet names aren't equal
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    if not sheets1 == sheets2:
        # print('')
        return False

    diff_dfs = {}
    # Check column labels and data for mismatches between sheets
    for sheet in sheets1:
        df1 = pd.DataFrame(wb1[sheet].values)
        df2 = pd.DataFrame(wb2[sheet].values)

        if not df1.equals(df2):
            comparison_values = df1.values == df2.values

            rows, cols = np.where(comparison_values == False)
            for item in zip(rows, cols):
                df1.iloc[item[0], item[1]] = '{} --> {}'.format(df1.iloc[item[0],
                                                                         item[1]],
                                                                df2.iloc[item[0],
                                                                         item[1]])

            diff_dfs.update({sheet: utils.first_row_to_cols(df1)})

    if diff_dfs:
        utils.write_report(file=diff_filename, report_dict=diff_dfs)
        return False
    else:
        return True


# Create a helper function for compare_workbooks_true() tests, parameterize for reports
# Delete diff files if they already exist
# Display warnings?

# Compare copies of the same workbook
def test_compare_workbooks_true():
    diff = ACTUAL_OUTPUTS_DIR + 'wb_1_wb_1_diff.xlsx'
    result = compare_workbooks(xlsx1=TEST_INPUTS_DIR + 'test_wb_1.xlsx',
                               xlsx2=EXPECTED_OUTPUTS_DIR + 'test_wb_1.xlsx',
                               diff_filename=diff)
    assert result is True
    # compare_workbooks() for test workbooks 1 and 1 should NOT export diff
    assert os.path.isfile(diff) is False


# Compare different workbooks
def test_compare_workbooks_false():
    # Compare workbooks with different data
    diff = ACTUAL_OUTPUTS_DIR + 'wb_1_wb_2_diff.xlsx'
    result_1_2 = compare_workbooks(xlsx1=TEST_INPUTS_DIR + 'test_wb_1.xlsx',
                                   xlsx2=EXPECTED_OUTPUTS_DIR + 'test_wb_2.xlsx',
                                   diff_filename=diff)
    assert result_1_2 is False
    # compare_workbooks() for test workbooks 1 and 2 should export diff
    assert os.path.isfile(diff) is True
    # Diff should contain all unequal sheets
    test_wb_1 = openpyxl.load_workbook(TEST_INPUTS_DIR + 'test_wb_1.xlsx')
    assert openpyxl.load_workbook(diff).sheetnames == test_wb_1.sheetnames

    # Compare workbooks with different sheet names
    diff = ACTUAL_OUTPUTS_DIR + 'wb_1_wb_3_diff.xlsx'
    result_1_3 = compare_workbooks(xlsx1=TEST_INPUTS_DIR + 'test_wb_1.xlsx',
                                   xlsx2=EXPECTED_OUTPUTS_DIR + 'test_wb_3.xlsx',
                                   diff_filename=diff)
    assert result_1_3 is False
    # compare_workbooks() for test workbooks 1 and 3 should NOT export diff
    assert os.path.isfile(diff) is False


def test_sites_report():
    sites_report.main(creds=creds,
                      sites_export=TEST_INPUTS_PEARS_DIR + "Site_Export.xlsx",
                      users_export=TEST_INPUTS_PEARS_DIR + "User_Export.xlsx",
                      output_dir=ACTUAL_OUTPUTS_DIR)
    report_filename = sites_report.report_filename()
    diff = ACTUAL_OUTPUTS_DIR + 'sites_report_diff.xlsx'
    result = compare_workbooks(xlsx1=ACTUAL_OUTPUTS_DIR + report_filename,
                               xlsx2=EXPECTED_OUTPUTS_DIR + report_filename,
                               diff_filename=diff)
    assert result is True
    # Report output changed if diff exists
    assert os.path.isfile(diff) is False


def test_staff_report():
    staff_report.main(creds=creds,
                      users_export=TEST_INPUTS_PEARS_DIR + "User_Export.xlsx",
                      program_activities_export=TEST_INPUTS_PEARS_DIR + "Program_Activities_Export.xlsx",
                      indirect_activities_export=TEST_INPUTS_PEARS_DIR + "Indirect_Activity_Export.xlsx",
                      coalitions_export=TEST_INPUTS_PEARS_DIR + "Coalition_Export.xlsx",
                      partnerships_export=TEST_INPUTS_PEARS_DIR + "Partnership_Export.xlsx",
                      pse_site_activities_export=TEST_INPUTS_PEARS_DIR + "PSE_Site_Activity_Export.xlsx",
                      success_stories_export=TEST_INPUTS_PEARS_DIR + "Success_Story_Export.xlsx",
                      staff_list=staff_list,
                      output_dir=ACTUAL_OUTPUTS_DIR)
    # CPHP Report
    report_filename_cphp = staff_report.report_filename(agency='CPHP')
    diff_cphp = ACTUAL_OUTPUTS_DIR + 'staff_report_cphp_diff.xlsx'
    result_cphp = compare_workbooks(xlsx1=ACTUAL_OUTPUTS_DIR + report_filename_cphp,
                                    xlsx2=EXPECTED_OUTPUTS_DIR + report_filename_cphp,
                                    diff_filename=diff_cphp)
    assert result_cphp is True
    assert os.path.isfile(diff_cphp) is False
    # SNAP-Ed Report
    report_filename_snap_ed = staff_report.report_filename(agency='SNAP-Ed')
    diff_snap_ed = ACTUAL_OUTPUTS_DIR + 'staff_report_snap_ed_diff.xlsx'
    result_snap_ed = compare_workbooks(xlsx1=ACTUAL_OUTPUTS_DIR + report_filename_snap_ed,
                                       xlsx2=EXPECTED_OUTPUTS_DIR + report_filename_snap_ed,
                                       diff_filename=diff_snap_ed)
    assert result_snap_ed is True
    assert os.path.isfile(diff_snap_ed) is False


def test_monthly_data_cleaning():
    monthly_data_cleaning.main(creds=creds,
                               coalitions_export=TEST_INPUTS_PEARS_DIR + "Coalition_Export.xlsx",
                               indirect_activities_export=TEST_INPUTS_PEARS_DIR + "Indirect_Activity_Export.xlsx",
                               partnerships_export=TEST_INPUTS_PEARS_DIR + "Partnership_Export.xlsx",
                               program_activities_export=TEST_INPUTS_PEARS_DIR + "Program_Activities_Export.xlsx",
                               pse_site_activities_export=TEST_INPUTS_PEARS_DIR + "PSE_Site_Activity_Export.xlsx",
                               staff_list=staff_list,
                               names_list=names_list,
                               unit_counties=unit_counties,
                               update_notifications=update_notifications,
                               output_dir=ACTUAL_OUTPUTS_DIR)
    report_filename = monthly_data_cleaning.report_filename(report='corrections')
    diff = ACTUAL_OUTPUTS_DIR + 'monthly_data_cleaning_diff.xlsx'
    result = compare_workbooks(xlsx1=ACTUAL_OUTPUTS_DIR + report_filename,
                               xlsx2=EXPECTED_OUTPUTS_DIR + report_filename,
                               diff_filename=diff)
    assert result is True
    assert os.path.isfile(diff) is False
